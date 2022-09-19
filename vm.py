from lib2to3.pytree import generate_matches
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side
from itertools import combinations


def main():
    try:
        # Only works for xl with xlsx ext
        wb = load_workbook('VMQuiz.xlsx')
    except Exception:
        wb = Workbook()

    # Worksheet
    ws = wb.active
    ws.title = 'VM Quiz 2022'

    # Group spacing
    group_col_space = 1
    first_groups_row = 2
    second_groups_row = 18

    # __First level groups__

    # Creating group A
    group_A_countries = ['Qatar', 'Ecuador', 'Senegal', 'Nederländerna']
    group_A_color = 'FF4500' # Orange Red
    group_A = Group('Group A', group_A_countries, first_groups_row, group_col_space, group_A_color, ws)
    
    # Creating group B
    group_B_countries = ['England', 'Iran', 'USA', 'Wales']
    group_B_color = '9ACD32' # yellow green
    group_B = Group('Group B', group_B_countries, first_groups_row, 2*group_col_space + 5, group_B_color, ws)

    # Creating group C
    group_C_countries = ['Argentina', 'Saudiarabien', 'Mexiko', 'Polen']
    group_C_color = 'FF8C00' # Dark Orange
    group_C = Group('Group C', group_C_countries, first_groups_row, 3*group_col_space + 10, group_C_color, ws)

    # Creating group D
    group_D_countries = ['Frankrike', 'Australien', 'Danmark', 'Tunisien']
    group_D_color = '1E90FF' # Dodger Blue
    group_D = Group('Group D', group_D_countries, first_groups_row, 4*group_col_space + 15, group_D_color, ws)

    # __Second level groups__

    # Creating group E
    group_E_countries = ['Spanien', 'Costa Rica', 'Tyskland', 'Japan']
    group_E_color = 'FFD700' # Gold
    group_E = Group('Group E', group_E_countries, second_groups_row, group_col_space, group_E_color, ws)

    # Creating group F
    group_F_countries = ['Belgien', 'Kanada', 'Marocko', 'Kroatien']
    group_F_color = 'A9A9A9' # Dark Grey
    group_F = Group('Group F', group_F_countries, second_groups_row, 2*group_col_space + 5, group_F_color, ws)

    # Creating group G
    group_G_countries = ['Brasilien', 'Serbien', 'Schweiz', 'Kamerun']
    group_G_color = 'EE82EE' # Violet
    group_G = Group('Group G', group_G_countries, second_groups_row, 3*group_col_space + 10, group_G_color, ws)

    # Creating group H
    group_H_countries = ['Portugal', 'Ghana', 'Uruguay', 'Sydkorea']
    group_H_color = '87CEEB' # Sky Blue
    group_H = Group('Group H', group_H_countries, second_groups_row, 4*group_col_space + 15, group_H_color, ws)
 

    # Set width of all the columns in range below
    for i in range(1,30):
        ws.column_dimensions[get_column_letter(i)].width = 15

    wb.save('VMQuiz.xlsx')



def set_value_to_cell(cell, value, color = None, alignment = None):

    # Borders
    regular = Side(border_style="thin", color="000000")
    border = Border(regular, regular, regular, regular)


    cell.value = value
    cell.border = border

    if color:
        cell.fill = color
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = Alignment(horizontal='left')


def get_cell(col, row):
    return get_column_letter(col) + str(row)

class Playoffs():

    def __init__(self, row_start, col_start):

        self.row_start = row_start
        self.col_start = col_start


        self.generate_playoff()

    def generate_playoff():
        pass

    def generate_16():
        pass

    def generate_quarterfinal():
        pass

    def generate_semifinal():
        pass

    def generate_final():
        pass

        

class Group():

    def __init__(self, group_name, countries, row_start, col_start, color, ws):
        self.group_name = group_name
        self.group_headers = [group_name, 'Points', 'GS', 'GC', 'GD']
        self.countries = countries
        self.row_start = row_start
        self.col_start = col_start
        self.col_start_offset = col_start + 6
        self.fill_color = PatternFill(patternType = 'solid', fgColor = color)
        self.ws = ws

        self.generate_group()

    def generate_group(self):
        
        self.generate_matches()

        self.generate_scoreboard()


    # Helper method to merge headers
    def merge_header(self, row, col, col_width, value):
        self.ws.merge_cells(f"{get_cell(col, row)}:{get_cell(col+col_width, row)}")
        cell = self.ws[get_cell(col, row)]

        set_value_to_cell(cell, value, self.fill_color)

    def generate_matches(self):
        row = self.row_start

        self.merge_header(row, self.col_start, 1, f"Matches {self.group_name}")
        self.merge_header(row, self.col_start+2, 1, "Score")
        cell = self.ws[get_cell(self.col_start + 4, row)]
        set_value_to_cell(cell, 'Result (1, X, 2)', self.fill_color)

        matches = list(combinations(self.countries, 2))
        for match in matches:
            row += 1
            for col_offset in range(5):
                cell = self.ws[get_cell(self.col_start + col_offset, row)]
                if col_offset == 0:
                    value = match[0]
                elif col_offset == 1:
                    value = match[1]
                elif col_offset == 4:
                    first_team_cell = get_cell(self.col_start + col_offset-2, row)
                    second_team_cell = get_cell(self.col_start + col_offset-1, row)
                    value = f'=IF({first_team_cell} > {second_team_cell}, 1, IF({second_team_cell} > {first_team_cell}, 2, "X"))'
                else:
                    value = ''
                
                set_value_to_cell(cell, value, self.fill_color)


    # Add country to group scoreboard
    def generate_scoreboard(self):

        row = self.row_start
        
        # Create scoreboard header
        for i in range(len(self.group_headers)):
            cell = self.ws[get_cell(self.col_start_offset + i, self.row_start)]
            set_value_to_cell(cell, self.group_headers[i], self.fill_color)

        # Create scoreboard body
        for r in range(len(self.countries)):
            row +=1
         
            for c in range(5):
                col_letter = get_column_letter(self.col_start_offset + c)
                pos = col_letter + str(row)
                cell = self.ws[pos]
                
                # Set country name
                if c == 0:
                    value = self.countries[r]
                # Get points
                elif c == 1:
                    value = self.get_points(r)
                # Get goal scored
                elif c == 2:
                    value = self.get_goals_scored(r)
                # Get goal conceded
                elif c == 3:
                    value = self.get_goals_conceded(r)
                # Get goal difference
                elif c == 4:
                    gs_cell = get_column_letter(self.col_start_offset + c-2) + str(row)
                    gc_cell = get_column_letter(self.col_start_offset + c-1) + str(row)
                    value = f"=SUM({gs_cell}, -{gc_cell})"

                set_value_to_cell(cell, value, self.fill_color)


    # Returns excel formula for getting the goals scored from the matches to the group scoreboard
    def get_goals_scored(self, i):

        c = self.col_start + 2
        r = self.row_start + 1

        if i == 0:
            return f'=SUM({get_cell(c, r)}, {get_cell(c, r+1)}, {get_cell(c, r+2)})' 
        elif i == 1:
            return f'=SUM({get_cell(c+1, r)}, {get_cell(c, r+3)}, {get_cell(c, r+4)})' 
        elif i == 2:
            return f'=SUM({get_cell(c+1, r+1)}, {get_cell(c+1, r+3)}, {get_cell(c, r+5)})' 
        elif i == 3:
            return f'=SUM({get_cell(c+1, r+2)}, {get_cell(c+1, r+4)}, {get_cell(c+1, r+5)})' 


    # Returns excel formula for getting the goals conceded from the matches to the group scoreboard
    def get_goals_conceded(self, i):
        
        c = self.col_start + 2
        r = self.row_start + 1

        if i == 0:
            return f'=SUM({get_cell(c+1, r)}, {get_cell(c+1, r+1)}, {get_cell(c+1, r+2)})' 
        elif i == 1:
            return f'=SUM({get_cell(c, r)}, {get_cell(c+1, r+3)}, {get_cell(c+1, r+4)})' 
        elif i == 2:
            return f'=SUM({get_cell(c, r+1)}, {get_cell(c, r+3)}, {get_cell(c+1, r+5)})' 
        elif i == 3:
            return f'=SUM({get_cell(c, r+2)}, {get_cell(c, r+4)}, {get_cell(c, r+5)})' 
        

    # Returns excel formula for getting the points from the matches to the group scoreboard
    def get_points(self, i):

        def points_from_result(result):

            def eval_result(result, invert):
                if invert:
                    return f'IF({result} = 2, 3, IF({result} = "X", 1, 0))'
                else:
                    return f'IF({result} = 1, 3, IF({result} = "X", 1, 0))'

            result_list = []
            for cell, invert in result.items():
                result_list.append(eval_result(cell, invert))
                
            return f'=SUM({result_list[0]}, {result_list[1]}, {result_list[2]})'

        c = self.col_start + 4
        r = self.row_start + 1

        if i == 0:
            result = {get_cell(c, r) : False, get_cell(c, r+1) : False, get_cell(c, r+2) : False}
        elif i == 1:
            result = {get_cell(c, r) : True, get_cell(c, r+3) : False, get_cell(c, r+4) : False}
        elif i == 2:
            result = {get_cell(c, r+1) : True, get_cell(c, r+3) : True, get_cell(c, r+5) : False}
        elif i == 3:
            result = {get_cell(c, r+2) : True, get_cell(c, r+4) : True, get_cell(c, r+5) : True}
        return points_from_result(result)
    
        
    

if __name__ == "__main__":
    main()