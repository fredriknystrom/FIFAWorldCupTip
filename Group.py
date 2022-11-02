from openpyxl.styles import PatternFill
from itertools import combinations
from util_funcs import get_cell, set_value_to_cell

class Group():

    def __init__(self, group_name, countries, color, row_start, col_start, ws):
        self.group_name = group_name
        self.group_headers = [group_name, 'Points', 'GS', 'GC', 'GD', 'Total']
        self.countries = countries
        self.row_start = row_start
        self.col_start = col_start
        self.col_start_offset = col_start + 6
        self.fill_color = PatternFill(patternType = 'solid', fgColor = color)
        self.ws = ws

        self.generate_group()

    def generate_group(self):
        
        self.generate_matches()

        self.generate_scoreboard()


    def generate_matches(self):
        row = self.row_start
        values = [f"Matches {self.group_name}", '', 'Score', '', 'Result (1, X, 2)']
        for i in range(5):
            cell = self.ws[get_cell(self.col_start+i, row)]
            set_value_to_cell(cell, values[i], self.fill_color)
        
        matches = list(combinations(self.countries, 2))
        for match in matches:
            row += 1
            for col_offset in range(5):
                cell = self.ws[get_cell(self.col_start + col_offset, row)]
                if col_offset == 0:
                    value = match[0]
                elif col_offset == 1:
                    value = match[1]
                elif col_offset == 4:
                    first_team_cell = get_cell(self.col_start + col_offset-2, row)
                    second_team_cell = get_cell(self.col_start + col_offset-1, row)
                    value = f'=IF({first_team_cell} > {second_team_cell}, 1, IF({second_team_cell} > {first_team_cell}, 2, "X"))'
                else:
                    value = ''
                
                set_value_to_cell(cell, value, self.fill_color)


    # Add country to group scoreboard
    def generate_scoreboard(self):

        row = self.row_start
        
        # Create scoreboard header
        for i in range(len(self.group_headers)):
            cell = self.ws[get_cell(self.col_start_offset + i, self.row_start)]
            set_value_to_cell(cell, self.group_headers[i], self.fill_color)

        # Create scoreboard body
        for r in range(len(self.countries)):
            row +=1
         
            for c in range(6):
                cell = self.ws[get_cell(self.col_start_offset + c, row)]
                
                # Set country name
                if c == 0:
                    value = self.countries[r]
                # Get points
                elif c == 1:
                    value = self.get_points(r)
                # Get goal scored
                elif c == 2:
                    value = self.get_goals_scored(r)
                # Get goal conceded
                elif c == 3:
                    value = self.get_goals_conceded(r)
                # Get goal difference
                elif c == 4:
                    gs_cell = get_cell(self.col_start_offset + 2, row)
                    gc_cell = get_cell(self.col_start_offset + 3, row)
                    value = f"=SUM({gs_cell}, -{gc_cell})"
                elif c == 5:
                    points_cell = get_cell(self.col_start_offset + 1, row)
                    gd_cell = get_cell(self.col_start_offset + 4, row)
                    value = f"=SUM({points_cell}*1000, {gd_cell}, {gs_cell}*0.001)"

                set_value_to_cell(cell, value, self.fill_color)


    # Returns excel formula for getting the goals scored from the matches to the group scoreboard
    def get_goals_scored(self, i):

        c = self.col_start + 2
        r = self.row_start + 1

        if i == 0:
            return f'=SUM({get_cell(c, r)}, {get_cell(c, r+1)}, {get_cell(c, r+2)})' 
        elif i == 1:
            return f'=SUM({get_cell(c+1, r)}, {get_cell(c, r+3)}, {get_cell(c, r+4)})' 
        elif i == 2:
            return f'=SUM({get_cell(c+1, r+1)}, {get_cell(c+1, r+3)}, {get_cell(c, r+5)})' 
        elif i == 3:
            return f'=SUM({get_cell(c+1, r+2)}, {get_cell(c+1, r+4)}, {get_cell(c+1, r+5)})' 


    # Returns excel formula for getting the goals conceded from the matches to the group scoreboard
    def get_goals_conceded(self, i):
        
        c = self.col_start + 2
        r = self.row_start + 1

        if i == 0:
            return f'=SUM({get_cell(c+1, r)}, {get_cell(c+1, r+1)}, {get_cell(c+1, r+2)})' 
        elif i == 1:
            return f'=SUM({get_cell(c, r)}, {get_cell(c+1, r+3)}, {get_cell(c+1, r+4)})' 
        elif i == 2:
            return f'=SUM({get_cell(c, r+1)}, {get_cell(c, r+3)}, {get_cell(c+1, r+5)})' 
        elif i == 3:
            return f'=SUM({get_cell(c, r+2)}, {get_cell(c, r+4)}, {get_cell(c, r+5)})' 
        

    # Returns excel formula for getting the points from the matches to the group scoreboard
    def get_points(self, i):

        def points_from_result(result):

            def eval_result(result, invert):
                if invert:
                    return f'IF({result} = 2, 3, IF({result} = "X", 1, 0))'
                else:
                    return f'IF({result} = 1, 3, IF({result} = "X", 1, 0))'

            result_list = []
            for cell, invert in result.items():
                result_list.append(eval_result(cell, invert))
                
            return f'=SUM({result_list[0]}, {result_list[1]}, {result_list[2]})'

        c = self.col_start + 4
        r = self.row_start + 1

        if i == 0:
            result = {get_cell(c, r) : False, get_cell(c, r+1) : False, get_cell(c, r+2) : False}
        elif i == 1:
            result = {get_cell(c, r) : True, get_cell(c, r+3) : False, get_cell(c, r+4) : False}
        elif i == 2:
            result = {get_cell(c, r+1) : True, get_cell(c, r+3) : True, get_cell(c, r+5) : False}
        elif i == 3:
            result = {get_cell(c, r+2) : True, get_cell(c, r+4) : True, get_cell(c, r+5) : True}
        return points_from_result(result)
        

    def get_winner(self):

        range1 = f'{get_cell(self.col_start_offset+5, self.row_start+1)}:{get_cell(self.col_start_offset+5, self.row_start+4)}'

        range2 = f'{get_cell(self.col_start_offset, self.row_start+1)}:{get_cell(self.col_start_offset, self.row_start+4)}'

        test = f'=INDEX({range2},MATCH(LARGE({range1},1),{range1},0))'

        return test



    def get_second(self):

        range1 = f'{get_cell(self.col_start_offset+5, self.row_start+1)}:{get_cell(self.col_start_offset+5, self.row_start+4)}'

        range2 = f'{get_cell(self.col_start_offset, self.row_start+1)}:{get_cell(self.col_start_offset, self.row_start+4)}'

        test = f'=INDEX({range2},MATCH(LARGE({range1},2),{range1},0))'

        return test
        

    def __repr__(self) -> str:
        return self.group_name