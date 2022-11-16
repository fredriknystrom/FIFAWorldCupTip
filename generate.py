
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from Group import Group
from Playoff import Playoff
import os


def main():
    path = os.path.abspath('quizes/troll.xlsx')
    if os.path.exists(path):
        os.remove(path)

    # Create workbook and activate worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'VM Quiz 2022'

    group_countires = [
        ['Qatar', 'Ecuador', 'Senegal', 'Nederländerna'],
        ['England', 'Iran', 'USA', 'Wales'],
        ['Argentina', 'Saudiarabien', 'Mexiko', 'Polen'],
        ['Frankrike', 'Australien', 'Danmark', 'Tunisien'],
        ['Spanien', 'Costa Rica', 'Tyskland', 'Japan'],
        ['Belgien', 'Kanada', 'Marocko', 'Kroatien'],
        ['Brasilien', 'Serbien', 'Schweiz', 'Kamerun'],
        ['Portugal', 'Ghana', 'Uruguay', 'Sydkorea']
    ]

    group_colors = [
        'FB4D3D',          
        '9ACD32',          
        'F7C548',          
        '1E90FF', 
        '226F54', 
        'FF8811',         
        'CB769E',          
        '7CDEDC',          
    ]

    playoff_color = '8D99AE'

    group_labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    # Group spacing
    col_start = 1
    row_start = 2
    row_offset = 8
 
    groups = []

    # Creating groups
    for i in range(len(group_countires)):
        groups.append(Group(f'Group {group_labels[i]}', group_countires[i],  group_colors[i], row_start +  i*row_offset , col_start, ws))

    Playoff(row_start, 14, groups, playoff_color, ws)


    
 
    # Set width of all the columns in range below
    for i in range(1,40):
        ws.column_dimensions[get_column_letter(i)].width = 15

    wb.save('quizes/empty_tip.xlsx')


if __name__ == "__main__":
    main()