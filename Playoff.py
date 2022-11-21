from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from util_funcs import get_cell, set_value_to_cell


class Playoff():

    def __init__(self, row_start, col_start, groups, color, ws):

        self.row_start = row_start
        self.col_start = col_start
        self.col_offset = 5
        self.fill_color = PatternFill(patternType = 'solid', fgColor = color)
        self.groups = groups
        self.ws = ws

        self.generate_playoff()


    def generate_playoff(self):
        self.generate_16()
        self.generate_quarterfinals()
        self.generate_semifinals()
        self.generate_bronze()
        self.generate_gold()
        self.generate_gold_bronze_and_top_scorer()


    def generate_16(self):
        cell = self.ws[get_cell(self.col_start, self.row_start)]
        set_value_to_cell(cell, 'Round of 16', self.fill_color)
        row = self.row_start + 1

        for r in range(0,8,2):
        
            text1 = f'Group {get_column_letter(r+1)}'
            text2 = f'Group {get_column_letter(r+2)}'

            for i in range(2):
                if i == 0:
                    first_group = self.groups[r]
                    second_group = self.groups[r+1]
                if i == 1:
                    first_group = self.groups[r+1]
                    second_group = self.groups[r]

                winner = first_group.get_winner()
                second = second_group.get_second()
     
                for c in range(4):
                    cell1 = self.ws[get_cell(self.col_start + c, row)]
                    cell2 = self.ws[get_cell(self.col_start + c, row+1)]
                    if c == 0:
                        value1 = f'1st in {text1}'
                        value2 = winner
                    if c == 1:
                        value1 = f'2nd in {text2}'
                        value2 = second
                    if c == 2:
                        value1 = 'Score'
                        value2 = ''
                    if c == 3:
                        value1 = ''
                        value2 = ''
                    set_value_to_cell(cell1, value1, self.fill_color)
                    set_value_to_cell(cell2, value2, self.fill_color)

                tmp = text1
                text1 = text2
                text2 = tmp
                row += 4


    def generate_finals(self, final_type, winner_func, col_offset):
        col = self.col_start + col_offset
        row = self.row_start

        if final_type not in ['Gold', 'Bronze']:
            cell = self.ws[get_cell(col, row)]
            set_value_to_cell(cell, final_type + 's', self.fill_color)
        row += 1
        if final_type == 'Bronze':
            row += 4

        teams = winner_func
        for r in range(0, len(teams), 2):

            for c in range(4):
                cell1 = self.ws[get_cell(col + c, row)]
                cell2 = self.ws[get_cell(col + c, row+1)]
                if c == 0:
                    if final_type == 'Gold' or 'Bronze':
                        value1 = final_type + ' Match'
                    else:
                        value1 = f'{final_type} {int(r/2+1)}'
                    value2 = teams[r]
                if c == 1:
                    value1 = f''
                    value2 = teams[r+1]
                if c == 2:
                    value1 = 'Score'
                    value2 = ''
                if c == 3:
                    value1 = ''
                    value2 = ''
                set_value_to_cell(cell1, value1, self.fill_color)
                set_value_to_cell(cell2, value2, self.fill_color)
            # adds space between matches
            row += 4


    def generate_quarterfinals(self):
        order = [20, 28, 4, 12, 24, 32, 16, 8]
        winners = self.get_round_of_16_winners(order)
        self.generate_finals('Quarterfinal', winners, self.col_offset)


    def generate_semifinals(self):
        order = [4, 8, 12, 16]
        winners = self.get_quaterfinal_winners(order)
        self.generate_finals('Semifinal', winners, 2*self.col_offset)

    def generate_bronze(self):
        order = [4, 8]
        losers = self.get_losers(order, 2*self.col_offset)
        self.generate_finals('Bronze', losers, 3*self.col_offset)


    def generate_gold(self):
        order = [4, 8]
        winners = self.get_semifinal_winners(order)
        self.generate_finals('Gold', winners, 3*self.col_offset)

    def generate_gold_bronze_and_top_scorer(self):
        row = self.row_start + 9
        col = self.col_start + 3*self.col_offset
        matrix = [['Gold Winner', 'Bronze Winner', 'Top Scorer', 'Goals Scored'],
                  [self.get_winner([4]), self.get_winner([8]),'', '']]
        for list in matrix:
            for value in list:
                if value != 'x':
                    cell = self.ws[get_cell(col, row)]
                    set_value_to_cell(cell, value, self.fill_color)
                col += 1
            row +=1
            col = self.col_start + 3*self.col_offset

    def get_winners(self, rows, col_offset):
        col = self.col_start + col_offset
        winners = []
       
        for row in rows:
            formula = f'=IF({get_cell(col+2, row)} > {get_cell(col+3, row)}, {get_cell(col, row)}, {get_cell(col+1, row)})'
            winners.append(formula)
        return winners

    def get_losers(self, rows, col_offset):
        col = self.col_start + col_offset
        winners = []
       
        for row in rows:
            formula = f'=IF({get_cell(col+2, row)} < {get_cell(col+3, row)}, {get_cell(col, row)}, {get_cell(col+1, row)})'
            winners.append(formula)
        return winners


    def get_round_of_16_winners(self, rows):
        return self.get_winners(rows, 0)
    

    def get_quaterfinal_winners(self, rows):
        return self.get_winners(rows, self.col_offset)


    def get_semifinal_winners(self, rows):
        return self.get_winners(rows, 2*self.col_offset)

    def get_winner(self, rows):
        return self.get_winners(rows, 3*self.col_offset)[0]