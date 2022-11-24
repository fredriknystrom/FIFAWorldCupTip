from openpyxl import load_workbook
from util_funcs import get_cell
import os
import operator
import codecs


def main():
    results_from_folder('chalmers', 'chalmers.txt')
    results_from_folder('family', 'family.txt')
    
def results_from_folder(folder_path, result_file_name):
    solution_wb = load_workbook('solution/solution.xlsx', data_only=True)
    solution_ws = solution_wb.active
    results = dict()

    for file in os.listdir(folder_path):
        if file != '.DS_Store':
            wb = load_workbook(f'{folder_path}/{file}', data_only=True)
            ws = wb.active
    
            name = file.split('.')[0]

            results[name] = compare_tip(ws, solution_ws)

    # sorts dict by values
    results = dict( sorted(results.items(), key=operator.itemgetter(1),reverse=True))

    with codecs.open(result_file_name, encoding='utf-8', mode='w') as result_file:
        num = 1
        for key, value in results.items():
            result_file.write(f'{num}.    {key}: {value} points <br>\n')
            num += 1


def compare_tip(ws, solution_ws):
    total_points = 0
    # one point per correct scored goals and one point if correct match result
    total_points += group_points(ws, solution_ws, [3, 4, 5])
    # one point per correct scored goals and two points per correct team into round of 16 
    #total_points += round_of_16_points(ws, solution_ws, 8, 2, [14, 15, 16, 17]) 
    # one point per correct scored goals and four points per correct team into quarterfinals
    #total_points += quarter_points(ws, solution_ws, 4, 4, [19, 20, 21, 22]) 
    # one point per correct scored goals and eight points per correct team into semifinals
    #total_points += semi_points(ws, solution_ws, 2, 8, [24, 25, 26, 27])
    # one point per correct scored goals and sixteen points per correct team into final
    #total_points += final_points(ws, solution_ws, 1, 16, [29, 30, 31, 32])
    # bronze match points
    #total_points += get_playoffs_points(ws, solution_ws, 1, 0, [29, 30, 31, 32], 4)
    # 16 points for correct winner
    #total_points += bronze_points(ws, solution_ws, 'AD12')
    # 32 points for correct winner
    #total_points += gold_points(ws, solution_ws, 'AC12')
    # ten points for top scorer and 16 points for correct number of goals
    #total_points += top_scorer_and_goals_points(ws, solution_ws, ['AE12', 'AF12'])

    return total_points


def group_points(ws, solution_ws, col_range):
    points = 0
    row = 3
    last_match_row = 64
    for i in range(3, last_match_row+1):
        if row in [9, 17, 25, 33, 41, 49, 57]:
            row += 2
        else:
            for col in col_range: # C, D, E
                cell = get_cell(col, row)
                if ws[cell].value == solution_ws[cell].value:
                    points += 1
            row += 1
        if row > last_match_row:
            return points

# get teams from solution worksheet from playoff
def get_teams(solution_ws, n_matches, col_range):
    teams = []
    for row in range(4, n_matches*4+1, 4):
        for col in col_range:
            cell = get_cell(col, row)
            teams.append(solution_ws[cell].value)
    return teams

# helper function to get points from different parts of the playoff
def get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range, row_offset=0):
    teams = get_teams(solution_ws, n_matches, col_range[0:3])
    points = 0
    for row in range(4+row_offset, n_matches*4+1+row_offset, 4):
        for col in col_range[0:2]:
            cell = get_cell(col, row)
            if ws[cell].value in teams:
                points += team_points
        for col in col_range[2:4]:
            cell = get_cell(col, row)
            if ws[cell].value == solution_ws[cell].value:
                points += 1
    return points


def round_of_16_points(ws, solution_ws, n_matches, team_points, col_range):
    return get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range)
    

def quarter_points(ws, solution_ws, n_matches, team_points, col_range):
    return get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range)


def semi_points(ws, solution_ws, n_matches, team_points, col_range):
    return get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range)


def final_points(ws, solution_ws, n_matches, team_points, col_range):
    return get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range)


def bronze_points(ws, solution_ws, cell):
    if ws[cell].value == solution_ws[cell].value:
        return 16
    else:
        return 0

def gold_points(ws, solution_ws, cell):
    if ws[cell].value == solution_ws[cell].value:
        return 32
    else:
        return 0


def top_scorer_and_goals_points(ws, solution_ws, cells):
    points = 0
    try:
        if ws[cells[0]].value.lower() == solution_ws[cells[0]].value.lower():
            points += 16
    except:
        print("No top scorer was filled in")
    finally:
        if ws[cells[1]].value == solution_ws[cells[1]].value:
            points += 16
        return points
    

if __name__ == '__main__':
    main()